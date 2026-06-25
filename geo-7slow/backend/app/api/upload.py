"""文件上传端点"""
import os
import re
import uuid
import zipfile
import tempfile
import shutil
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, Form, Query, HTTPException

from app.config import UPLOAD_DIR, UPLOAD_SLOTS, SLOT_MATCH_RULES, BAND_CONFLICTS
from app.models.schemas import (
    FileMeta, UploadResponse, ZipUploadResponse,
    SlotMatch, SlotInfo, SlotMatchError,
)

router = APIRouter(prefix="/api", tags=["upload"])


def _validate_geotiff(path: Path) -> dict | None:
    """验证GeoTIFF文件并返回元数据"""
    try:
        import rasterio
        with rasterio.open(str(path)) as src:
            bounds = list(src.bounds)
            res_x = abs(src.transform.a)
            res_y = abs(src.transform.e)
            return {
                "crs": str(src.crs) if src.crs else None,
                "resolution": round((res_x + res_y) / 2, 2),
                "width": src.width,
                "height": src.height,
                "bounds": [round(b, 6) for b in bounds],
            }
    except Exception:
        return None


def _excel_to_kml(excel_path: Path, kml_output_path: Path) -> dict | None:
    """读取 Excel 经纬度坐标，生成 KML 文件。返回 bounds 信息。

    自动识别经纬度列：经度范围 -180~180，纬度范围 -90~90。
    支持 2 列或 3+ 列（自动跳过序号列）。
    """
    try:
        from openpyxl import load_workbook
        from shapely.geometry import Polygon

        wb = load_workbook(str(excel_path), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return None

        # 判断第一行是否为表头（非数字）
        start_idx = 0
        if rows[0]:
            try:
                [float(v) for v in rows[0] if v is not None]
            except (ValueError, TypeError):
                start_idx = 1

        # 收集所有数字值
        num_cols = 0
        data_rows = []
        for row in rows[start_idx:]:
            nums = []
            for v in (row or []):
                try:
                    nums.append(float(v))
                except (ValueError, TypeError):
                    pass
            if len(nums) >= 2:
                data_rows.append(nums)
                num_cols = max(num_cols, len(nums))

        if len(data_rows) < 3:
            return None

        # 自动识别经度列和纬度列
        # 取所有列的值范围
        col_ranges = {}
        for col_i in range(num_cols):
            vals = [r[col_i] for r in data_rows if col_i < len(r)]
            if vals:
                col_ranges[col_i] = (min(vals), max(vals))

        lon_col = None
        lat_col = None
        for col_i, (vmin, vmax) in col_ranges.items():
            if -180 <= vmin and vmax <= 180 and (vmax - vmin) > 0.01:
                if -90 <= vmin and vmax <= 90:
                    # 可同时是经度或纬度，先标记
                    if lon_col is None:
                        lon_col = col_i
                    elif lat_col is None:
                        lat_col = col_i
                else:
                    # 超出 ±90 范围，一定是经度
                    lon_col = col_i

        # 如果还没分清，用范围判断：范围更大的那列是经度（中国经度跨度大）
        if lon_col is not None and lat_col is None:
            # 只找到一个在 ±180 内的列，看另一列
            for col_i, (vmin, vmax) in col_ranges.items():
                if col_i != lon_col and -90 <= vmin and vmax <= 90:
                    lat_col = col_i
                    break

        if lon_col is None or lat_col is None:
            # 兜底：取最后两列（跳过序号列），左经右纬
            lon_col = max(0, num_cols - 2)
            lat_col = max(1, num_cols - 1)

        # 如果 lon_col > lat_col，说明序号在前，经纬度在后面
        # 确保经度值确实合理
        lon_vals = [col_ranges.get(lon_col, (0, 0))]
        lat_vals = [col_ranges.get(lat_col, (0, 0))]

        # 交换检查：如果"经度列"范围在 ±90 内而"纬度列"超出，说明搞反了
        if (col_ranges.get(lon_col, (0, 0))[1] <= 90 and
            col_ranges.get(lon_col, (0, 0))[0] >= -90 and
            col_ranges.get(lat_col, (0, 0))[1] > 90):
            lon_col, lat_col = lat_col, lon_col

        coords = []
        for row in data_rows:
            lon = row[lon_col] if lon_col < len(row) else None
            lat = row[lat_col] if lat_col < len(row) else None
            if lon is not None and lat is not None:
                coords.append((lon, lat))

        if len(coords) < 3:
            return None

        # 构建多边形并生成 KML
        poly = Polygon(coords)
        if not poly.is_valid:
            poly = poly.buffer(0)
        bounds = [round(b, 6) for b in poly.bounds]

        coord_str = " ".join(f"{lon},{lat},0" for lon, lat in coords)
        if coords[0] != coords[-1]:
            coord_str += f" {coords[0][0]},{coords[0][1]},0"

        kml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>Study Area</name>
    <Placemark>
      <Polygon>
        <outerBoundaryIs>
          <LinearRing>
            <coordinates>{coord_str}</coordinates>
          </LinearRing>
        </outerBoundaryIs>
      </Polygon>
    </Placemark>
  </Document>
</kml>"""
        kml_output_path.write_text(kml_content, encoding="utf-8")
        return {"crs": "EPSG:4326", "bounds": bounds, "feature_count": 1}
    except Exception as e:
        print(f"[excel_to_kml] ERROR: {e}")
        import traceback; traceback.print_exc()
        return None


def _validate_kml(path: Path) -> dict | None:
    """验证KML文件"""
    try:
        import fiona
        fiona.drvsupport.supported_drivers["KML"] = "rw"
        with fiona.open(str(path), driver="KML") as src:
            bounds = src.bounds
            return {
                "crs": str(src.crs) if src.crs else "EPSG:4326",
                "bounds": [round(b, 6) for b in bounds],
                "feature_count": len(list(src)),
            }
    except Exception:
        return None


@router.post("/upload", response_model=UploadResponse)
async def upload_files(
    dem: UploadFile = File(None),
    s2_b03: UploadFile = File(None),
    s2_b04: UploadFile = File(None),
    s2_b08: UploadFile = File(None),
    aster_b05: UploadFile = File(None),
    aster_b06: UploadFile = File(None),
    aster_b07: UploadFile = File(None),
    aster_b08: UploadFile = File(None),
    aster_b10: UploadFile = File(None),
    aster_b11: UploadFile = File(None),
    aster_b12: UploadFile = File(None),
    aster_b13: UploadFile = File(None),
    aster_b14: UploadFile = File(None),
    insar: UploadFile = File(None),
    kml: UploadFile = File(None),
):
    upload_id = uuid.uuid4().hex[:12]
    upload_dir = UPLOAD_DIR / upload_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    file_map = {
        "dem": dem, "s2_b03": s2_b03, "s2_b04": s2_b04, "s2_b08": s2_b08,
        "aster_b05": aster_b05, "aster_b06": aster_b06, "aster_b07": aster_b07,
        "aster_b08": aster_b08, "aster_b10": aster_b10, "aster_b11": aster_b11,
        "aster_b12": aster_b12, "aster_b13": aster_b13, "aster_b14": aster_b14,
        "insar": insar, "kml": kml,
    }

    saved_files = []
    for slot_name, upload_file in file_map.items():
        if upload_file is None or not upload_file.filename:
            if UPLOAD_SLOTS.get(slot_name, {}).get("required", False):
                raise HTTPException(400, f"缺少必填文件: {UPLOAD_SLOTS[slot_name]['label']}")
            continue

        ext = Path(upload_file.filename).suffix.lower()
        save_path = upload_dir / f"{slot_name}{ext}"

        content = await upload_file.read()
        save_path.write_bytes(content)

        meta = {"slot": slot_name, "filename": upload_file.filename}

        if slot_name == "kml":
            if ext in (".xlsx", ".xls"):
                # Excel 文件 → 转换为 KML
                kml_path = upload_dir / "kml.kml"
                info = _excel_to_kml(save_path, kml_path)
                if info is None:
                    raise HTTPException(400, f"Excel坐标文件无效（需要至少3个经纬度点）: {upload_file.filename}")
                save_path.unlink(missing_ok=True)  # 删除原始 Excel
                save_path = kml_path
            else:
                # .ovkml 复制为 .kml 后解析
                if ext == ".ovkml":
                    kml_path = upload_dir / "kml.kml"
                    shutil.copy2(str(save_path), str(kml_path))
                    save_path.unlink(missing_ok=True)
                    save_path = kml_path
                info = _validate_kml(save_path)
                if info is None:
                    raise HTTPException(400, f"KML/OVKML文件无效: {upload_file.filename}")
            meta.update(info)
        else:
            info = _validate_geotiff(save_path)
            if info is None:
                raise HTTPException(400, f"GeoTIFF文件无效: {upload_file.filename}")
            meta.update(info)

        saved_files.append(FileMeta(**meta))

    return UploadResponse(upload_id=upload_id, files=saved_files)


@router.get("/upload/{upload_id}/files")
async def list_uploaded_files(upload_id: str):
    """列出一个上传会话的所有文件"""
    upload_dir = UPLOAD_DIR / upload_id
    if not upload_dir.exists():
        raise HTTPException(404, "上传会话不存在")

    files = []
    for f in sorted(upload_dir.iterdir()):
        if f.is_file():
            slot = f.stem
            meta = {"slot": slot, "filename": f.name}
            if f.suffix.lower() == ".kml":
                info = _validate_kml(f)
            else:
                info = _validate_geotiff(f)
            if info:
                meta.update(info)
            files.append(meta)
    return {"upload_id": upload_id, "files": files}


# ─── ZIP 自动匹配逻辑 ────────────────────────────────────

def _get_context(entry_path: str) -> set[str]:
    """从文件完整路径（含目录）中提取上下文关键词"""
    parts = entry_path.lower().replace("\\", "/").split("/")
    keywords = set()
    for part in parts:
        for kw in ["sentinel", "s2", "msi", "aster", "swir", "tir"]:
            if kw in part:
                keywords.add(kw)
    return keywords


def match_files_to_slots(entries: list[str]) -> tuple[dict, list[str]]:
    """
    将 ZIP 内文件条目匹配到 slot。
    返回 (slot_key -> original_filename, warnings)
    """
    # 步骤 1: 过滤无效条目
    valid_exts = {".tif", ".tiff", ".kml", ".ovkml", ".xlsx", ".xls"}
    filtered = []
    for entry in entries:
        if entry.endswith("/"):
            continue
        basename = Path(entry).name
        if basename.startswith(".") or basename.startswith("_"):
            continue
        if "__macosx" in entry.lower():
            continue
        ext = Path(entry).suffix.lower()
        if ext in valid_exts:
            filtered.append(entry)

    matched = {}
    warnings = []
    unmatched_files = list(filtered)

    # 步骤 2: 匹配 KML/Excel 边界文件（无歧义）
    boundary_files = [e for e in unmatched_files if Path(e).suffix.lower() in (".kml", ".ovkml", ".xlsx", ".xls")]
    if boundary_files:
        # 优先用 KML/OVKML，其次 Excel
        kml_first = [f for f in boundary_files if Path(f).suffix.lower() in (".kml", ".ovkml")]
        chosen = kml_first[0] if kml_first else boundary_files[0]
        matched["kml"] = chosen
        unmatched_files.remove(chosen)
        if len(boundary_files) > 1:
            warnings.append(f"发现多个边界文件，使用: {Path(chosen).name}")

    # 步骤 3: 匹配 DEM（无歧义）
    dem_rule = SLOT_MATCH_RULES["dem"]
    for entry in list(unmatched_files):
        name_lower = Path(entry).name.lower()
        if Path(entry).suffix.lower() not in dem_rule["extensions"]:
            continue
        if any(p in name_lower for p in dem_rule["patterns"]):
            matched["dem"] = entry
            unmatched_files.remove(entry)
            break

    # 步骤 4: 匹配 InSAR（无歧义）
    insar_rule = SLOT_MATCH_RULES["insar"]
    for entry in list(unmatched_files):
        name_lower = Path(entry).name.lower()
        if Path(entry).suffix.lower() not in insar_rule["extensions"]:
            continue
        if any(p in name_lower for p in insar_rule["patterns"]):
            matched["insar"] = entry
            unmatched_files.remove(entry)
            break

    # 步骤 4.5: 匹配 InSAR 相干性（无歧义，与速度场配套）
    coh_rule = SLOT_MATCH_RULES["insar_coherence"]
    for entry in list(unmatched_files):
        name_lower = Path(entry).name.lower()
        if Path(entry).suffix.lower() not in coh_rule["extensions"]:
            continue
        if any(p in name_lower for p in coh_rule["patterns"]):
            matched["insar_coherence"] = entry
            unmatched_files.remove(entry)
            break

    # 步骤 5: 匹配波段文件（可能歧义）
    # 先收集所有带波段号的文件（支持 B8 和 B08 两种写法）
    band_pattern = re.compile(r"b(\d{1,2})(?:\b|_)", re.IGNORECASE)
    band_files = {}  # band_number -> [(entry, context)]
    for entry in list(unmatched_files):
        ext = Path(entry).suffix.lower()
        if ext not in {".tif", ".tiff"}:
            continue
        name_lower = Path(entry).name.lower()
        m = band_pattern.search(name_lower)
        if m:
            band_num = m.group(1).zfill(2)  # "8" → "08", "05" → "05"
            ctx = _get_context(entry)
            band_files.setdefault(band_num, []).append((entry, ctx))

    # 步骤 6: 对每个波段号分配 slot
    for band_num, candidates in band_files.items():
        # 查找冲突组
        conflict_slots = BAND_CONFLICTS.get(f"b{band_num}")

        if conflict_slots:
            # 有歧义的波段号（如 b08）
            # 先按上下文分组
            s2_cands = []
            aster_cands = []
            neutral_cands = []
            for entry, ctx in candidates:
                if any(k in ctx for k in ["sentinel", "s2", "msi"]):
                    s2_cands.append(entry)
                elif any(k in ctx for k in ["aster", "swir", "tir"]):
                    aster_cands.append(entry)
                else:
                    neutral_cands.append(entry)

            # 有上下文的先分配
            for entry in s2_cands:
                if "s2_b08" not in matched:
                    matched["s2_b08"] = entry
                    if entry in unmatched_files:
                        unmatched_files.remove(entry)
            for entry in aster_cands:
                if "aster_b08" not in matched:
                    matched["aster_b08"] = entry
                    if entry in unmatched_files:
                        unmatched_files.remove(entry)

            # 无上下文的：如果 S2 已有数据则优先给 ASTER，反之亦然
            # 判断逻辑：ASTER b05-b07 都匹配了 → 有 ASTER 数据 → B08 优先给 ASTER
            aster_context_present = all(
                f"aster_b{k}" in matched for k in ["05", "06", "07"]
            )
            # 同时看 S2 b03/b04 是否匹配了
            s2_context_present = all(
                f"s2_b{k}" in matched for k in ["03", "04"]
            )
            for entry in neutral_cands:
                if aster_context_present and not s2_context_present and "aster_b08" not in matched:
                    matched["aster_b08"] = entry
                elif s2_context_present and not aster_context_present and "s2_b08" not in matched:
                    matched["s2_b08"] = entry
                elif aster_context_present and "aster_b08" not in matched:
                    matched["aster_b08"] = entry
                elif "s2_b08" not in matched:
                    matched["s2_b08"] = entry
                elif "aster_b08" not in matched:
                    matched["aster_b08"] = entry
                else:
                    warnings.append(
                        f"波段 B{band_num} 已有匹配，跳过: {Path(entry).name}"
                    )
                    continue
                if entry in unmatched_files:
                    unmatched_files.remove(entry)
        else:
            # 无歧义的波段号，查找对应的 slot
            slot_key = None
            for sk, rule in SLOT_MATCH_RULES.items():
                if f"b{band_num}" in rule.get("patterns", []):
                    slot_key = sk
                    break

            if slot_key:
                for entry, ctx in candidates:
                    if entry in unmatched_files:
                        matched[slot_key] = entry
                        unmatched_files.remove(entry)
                        break

    # 步骤 7: 语义别名匹配
    for entry in list(unmatched_files):
        name_lower = Path(entry).name.lower().replace("-", "_").replace(" ", "_")
        for slot_key, rule in SLOT_MATCH_RULES.items():
            if slot_key in matched:
                continue
            aliases = rule.get("aliases", [])
            if any(alias in name_lower for alias in aliases):
                matched[slot_key] = entry
                unmatched_files.remove(entry)
                break

    return matched, warnings


def _build_upload_status(upload_id: str) -> ZipUploadResponse:
    """扫描上传目录，构建当前匹配状态"""
    upload_dir = UPLOAD_DIR / upload_id
    if not upload_dir.exists():
        # 目录不存在时返回空状态
        all_slots = [
            SlotInfo(slot=k, label=v["label"], required=v["required"])
            for k, v in UPLOAD_SLOTS.items()
        ]
        required_missing = [k for k, v in UPLOAD_SLOTS.items() if v["required"]]
        return ZipUploadResponse(
            upload_id=upload_id,
            matched=[], unmatched=all_slots, errors=[],
            required_missing=required_missing,
            all_required_filled=False,
        )

    matched = []
    unmatched = []
    errors = []
    required_missing = []

    for slot_key, slot_info in UPLOAD_SLOTS.items():
        # 查找该 slot 的文件
        slot_file = None
        for ext in [".tif", ".tiff", ".kml", ".ovkml"]:
            p = upload_dir / f"{slot_key}{ext}"
            if p.exists():
                slot_file = p
                break

        if slot_file:
            meta_dict = {"slot": slot_key, "filename": slot_file.name}
            if slot_key == "kml":
                info = _validate_kml(slot_file)
            else:
                info = _validate_geotiff(slot_file)

            if info:
                meta_dict.update(info)
                matched.append(SlotMatch(
                    slot=slot_key,
                    original_filename=slot_file.name,
                    meta=FileMeta(**meta_dict),
                ))
            else:
                errors.append(SlotMatchError(
                    slot=slot_key,
                    original_filename=slot_file.name,
                    error="文件格式无效",
                ))
                if slot_info["required"]:
                    required_missing.append(slot_key)
        else:
            unmatched.append(SlotInfo(
                slot=slot_key,
                label=slot_info["label"],
                required=slot_info["required"],
            ))
            if slot_info["required"]:
                required_missing.append(slot_key)

    return ZipUploadResponse(
        upload_id=upload_id,
        matched=matched,
        unmatched=unmatched,
        errors=errors,
        required_missing=required_missing,
        all_required_filled=len(required_missing) == 0,
    )


@router.post("/upload/zip", response_model=ZipUploadResponse)
async def upload_zip(
    zipfile_upload: UploadFile = File(...),
    existing_id: str = Query(None, description="已有上传会话ID（如KML已上传）"),
):
    """上传 ZIP 压缩包，自动匹配文件到 slot"""
    if not zipfile_upload.filename or not zipfile_upload.filename.lower().endswith(".zip"):
        raise HTTPException(400, "请上传 .zip 格式的压缩包")

    upload_id = existing_id if existing_id else uuid.uuid4().hex[:12]
    upload_dir = UPLOAD_DIR / upload_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    # 保存 zip 到临时文件
    tmp_dir = upload_dir / "_tmp_extract"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    zip_path = tmp_dir / "upload.zip"
    content = await zipfile_upload.read()
    zip_path.write_bytes(content)

    try:
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            entries = zf.namelist()
            matched_map, warnings = match_files_to_slots(entries)

            # 提取匹配的文件
            for slot_key, entry_path in matched_map.items():
                ext = Path(entry_path).suffix.lower()
                target_path = upload_dir / f"{slot_key}{ext}"
                with zf.open(entry_path) as src, open(target_path, "wb") as dst:
                    shutil.copyfileobj(src, dst)

                # Excel 边界文件 → 转换为 KML
                if slot_key == "kml" and ext in (".xlsx", ".xls"):
                    kml_path = upload_dir / "kml.kml"
                    info = _excel_to_kml(target_path, kml_path)
                    if info is None:
                        target_path.unlink(missing_ok=True)
                        warnings.append(f"Excel 文件 {Path(entry_path).name} 坐标无效，已跳过")
                    else:
                        target_path.unlink(missing_ok=True)
                # .ovkml → 重命名为 .kml
                elif slot_key == "kml" and ext == ".ovkml":
                    kml_path = upload_dir / "kml.kml"
                    shutil.move(str(target_path), str(kml_path))
    except zipfile.BadZipFile:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise HTTPException(400, "ZIP 文件损坏或格式不正确")
    finally:
        # 清理临时目录
        shutil.rmtree(tmp_dir, ignore_errors=True)

    return _build_upload_status(upload_id)


@router.post("/upload/{upload_id}/supplement", response_model=ZipUploadResponse)
async def supplement_file(
    upload_id: str,
    slot: str = Query(..., description="slot 名称"),
    file: UploadFile = File(...),
):
    """补充上传单个文件到已有上传会话"""
    import traceback
    if slot not in UPLOAD_SLOTS:
        raise HTTPException(400, f"未知的 slot: {slot}")

    upload_dir = UPLOAD_DIR / upload_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    ext = Path(file.filename).suffix.lower() if file.filename else ".tif"
    print(f"[supplement] slot={slot}, filename={file.filename}, ext={ext}")
    save_path = upload_dir / f"{slot}{ext}"

    content = await file.read()
    save_path.write_bytes(content)

    # 验证
    if slot == "kml":
        if ext in (".xlsx", ".xls"):
            kml_path = upload_dir / "kml.kml"
            info = _excel_to_kml(save_path, kml_path)
            if info is None:
                save_path.unlink(missing_ok=True)
                raise HTTPException(400, "Excel坐标无效（需要至少3个经纬度点，第1列经度第2列纬度）")
            save_path.unlink(missing_ok=True)
            save_path = kml_path
        else:
            # .kml 或 .ovkml：如果是 .ovkml，复制为 .kml 后由 Fiona 解析
            if ext == ".ovkml":
                kml_path = upload_dir / "kml.kml"
                shutil.copy2(str(save_path), str(kml_path))
                save_path.unlink(missing_ok=True)
                save_path = kml_path
            info = _validate_kml(save_path)
            if info is None:
                save_path.unlink(missing_ok=True)
                raise HTTPException(400, "KML/OVKML 文件无效")
    else:
        info = _validate_geotiff(save_path)
        if info is None:
            save_path.unlink(missing_ok=True)
            raise HTTPException(400, "GeoTIFF 文件无效")

    return _build_upload_status(upload_id)


@router.get("/upload/{upload_id}/status", response_model=ZipUploadResponse)
async def get_upload_status(upload_id: str):
    """获取上传会话的当前匹配状态"""
    return _build_upload_status(upload_id)
